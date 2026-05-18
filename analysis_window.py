"""
Raman Data Viewer – Analysis Window
=====================================
Layout (main window)
--------------------
  ┌─────────────────────────────┬──────────────────────────┐
  │  Heatmap                    │  Spectrum (V-slices)     │
  │  draggable V / H lines      ├──────────────────────────┤
  │  draggable rect regions     │  Time traces (H-slices)  │
  ├─────────────────────────────┴──────────────────────────┤
  │  Echem panel  ── UV/VIS intensity vs time              │
  │               ── Potential vs time                     │
  │  Vertical slice / region boundaries mirrored as        │
  │  dashed marker lines on the echem time axis.           │
  └────────────────────────────────────────────────────────┘
  Bottom: slice bar / region bar  (always visible)

Performance notes
-----------------
* Heatmap drag uses matplotlib blit animation: on drag start the canvas
  background is captured (minus the slice/region artists), and each
  motion event restores that bitmap then redraws only the moving
  artists via ax.draw_artist() + canvas.blit().  No full rasterise
  occurs until the drag ends or the heatmap itself changes.

* Side-panel (spectrum/time) updates during drag are throttled to at
  most one redraw every SIDE_PANEL_THROTTLE_MS milliseconds so that
  fast mouse movements do not queue a backlog of expensive redraws.

* PlotPanel uses uid→Line2D caches and set_data() / set_xdata() /
  set_ydata() incremental updates — never cla().

Key classes
-----------
PlotPanel
    Owns one Figure/Axes in a Tk frame.  Maintains uid→Line2D so that
    updates call set_data()+relim instead of cla().  Also manages a
    separate uid→Line2D dict for time-marker synchronisation.

EchemData
    Parsed contents of one xlsx sheet: time, potential, uv arrays.

EchemPanel
    Two stacked PlotPanels (UV-vis + potential) with sync_markers().

RegionPlotWindow
    Floating Toplevel with averaged spectrum + time trace per Region.

xlsx sheet auto-selection
    Best sheet is chosen by SequenceMatcher ratio against the Raman
    CSV filename stem; user can override via Combobox.
"""

from __future__ import annotations

import os
import sys
import csv
import time
import traceback
import tkinter as tk
from tkinter import ttk, messagebox, colorchooser, filedialog
from typing import Optional
from difflib import SequenceMatcher
from dataclasses import dataclass

import numpy as np
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.patches import Rectangle
from matplotlib.lines import Line2D

import theme as T


# How many milliseconds must elapse between side-panel redraws during drag.
# Lower = more responsive but more CPU; 40 ms ≈ 25 fps which feels live.
SIDE_PANEL_THROTTLE_MS: int = 40


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _next_color(used: list[str]) -> str:
    for c in T.SLICE_COLORS:
        if c not in used:
            return c
    return T.SLICE_COLORS[len(used) % len(T.SLICE_COLORS)]


def _style_axes(ax, xlabel: str, ylabel: str, title: str,
                axes_facecolor: str = T.BG_AXES):
    ax.set_facecolor(axes_facecolor)
    ax.tick_params(colors=T.FG_TICK, labelsize=T.FONT_SIZE_SMALL)
    for sp in ax.spines.values():
        sp.set_edgecolor(T.COLOR_SPINE)
    ax.set_xlabel(xlabel, color=T.FG_AXIS_LABEL, fontsize=T.FONT_SIZE_SMALL)
    ax.set_ylabel(ylabel, color=T.FG_AXIS_LABEL, fontsize=T.FONT_SIZE_SMALL)
    ax.set_title(title, color=T.FG_PLOT_TITLE, fontsize=T.FONT_SIZE_BODY)
    ax.grid(True, color=T.COLOR_GRID, linewidth=0.5, linestyle="--", alpha=0.7)


def _similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


# ─────────────────────────────────────────────────────────────────────────────
# CSV loader (Raman)
# ─────────────────────────────────────────────────────────────────────────────

def _filter_positive_times(times: np.ndarray,
                            intensities: np.ndarray) -> tuple[np.ndarray, np.ndarray]:
    mask = times >= 0
    return times[mask], intensities[:, mask]


def load_raman_csv(path: str) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    """Return (wavenumbers [N], times [M], intensity [N×M])."""
    with open(path, newline="", encoding="utf-8-sig") as fh:
        sample = fh.read(4096)
    delim = "," if sample.count(",") >= sample.count(";") else ";"

    with open(path, newline="", encoding="utf-8-sig") as fh:
        rows = [r for r in csv.reader(fh, delimiter=delim)
                if any(c.strip() for c in r)]

    def _f(s: str) -> float:
        return float(s.strip().replace(",", "."))

    times = np.array([_f(c) for c in rows[0][1:] if c.strip()])
    wavenumbers, intensities = [], []
    for row in rows[1:]:
        if not row or not row[0].strip():
            continue
        try:
            wn = _f(row[0])
        except ValueError:
            continue
        vals = []
        for c in row[1: 1 + len(times)]:
            try:
                vals.append(_f(c))
            except ValueError:
                vals.append(np.nan)
        if vals:
            wavenumbers.append(wn)
            intensities.append(vals)

    wavenumbers = np.array(wavenumbers, dtype=float)
    intensities = np.array(intensities, dtype=float)
    if intensities.ndim == 2 and intensities.shape[1] != len(times):
        intensities = intensities[:, : len(times)]

    return wavenumbers, *_filter_positive_times(times, intensities)


# ─────────────────────────────────────────────────────────────────────────────
# xlsx loader (Echem)
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class EchemData:
    """Parsed electrochemistry data from one xlsx sheet."""
    sheet_name:      str
    time:            np.ndarray   # [M]  seconds
    potential:       np.ndarray   # [M]  V
    uv:              np.ndarray   # [M]  UV/VIS
    potential_label: str = "Potential [V]"
    uv_label:        str = "UV/VIS"


def _to_float_array(values: list) -> np.ndarray:
    out = []
    for v in values:
        if v is None or v == "":
            out.append(np.nan)
            continue
        try:
            out.append(float(str(v).replace(",", ".")))
        except (ValueError, TypeError):
            out.append(np.nan)
    return np.array(out, dtype=float)


def load_echem_xlsx(path: str) -> dict[str, EchemData]:
    """
    Load all sheets from *path*.  Returns sheet_name → EchemData.

    Column detection
    ----------------
    Sheets contain interleaved groups of (Time, <data>) column pairs.
    We locate:
      • potential column – header containing "DAQ" + "poten", or just "poten"
      • UV column        – header containing "UV"

    The Time column that immediately precedes each data column is used as
    its time axis.  If the two time axes differ the UV trace is interpolated
    onto the potential time axis.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required to load xlsx files.\n"
            "Install it with:  pip install openpyxl"
        )

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result: dict[str, EchemData] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows: list[tuple] = list(ws.iter_rows(values_only=True))
        if len(all_rows) < 2:
            continue

        # find first row with at least one non-empty string → header
        header_idx = 0
        for i, row in enumerate(all_rows):
            if any(isinstance(c, str) and c.strip() for c in row):
                header_idx = i
                break

        header    = [str(c).strip() if c is not None else "" for c in all_rows[header_idx]]
        data_rows = all_rows[header_idx + 1:]
        if not data_rows:
            continue

        # helper: find first column whose header contains ALL keywords
        def _find_col(keywords: list[str],
                      exclude: list[str] | None = None) -> int | None:
            kw = [k.lower() for k in keywords]
            ex = [e.lower() for e in (exclude or [])]
            for ci, h in enumerate(header):
                hl = h.lower()
                if all(k in hl for k in kw) and not any(e in hl for e in ex):
                    return ci
            return None

        time_cols = [i for i, h in enumerate(header) if h.lower() == "time"]

        def _time_for(data_col: int | None) -> int | None:
            if data_col is None:
                return None
            candidates = [t for t in time_cols if t <= data_col]
            return max(candidates) if candidates else (time_cols[0] if time_cols else None)

        pot_col = _find_col(["daq", "poten"]) or _find_col(["poten"])
        uv_col  = _find_col(["uv"])

        t_pot_col = _time_for(pot_col)
        t_uv_col  = _time_for(uv_col)

        def _col(idx: int | None) -> list:
            if idx is None:
                return []
            return [row[idx] if idx < len(row) else None for row in data_rows]

        t_pot_vals = _col(t_pot_col)
        pot_vals   = _col(pot_col)
        t_uv_vals  = _col(t_uv_col)
        uv_vals    = _col(uv_col)

        t_arr   = _to_float_array(t_pot_vals if t_pot_vals else t_uv_vals)
        pot_arr = _to_float_array(pot_vals)
        uv_arr  = _to_float_array(uv_vals)

        # interpolate UV onto potential time axis if they differ
        if t_uv_vals and t_pot_vals and t_uv_col != t_pot_col:
            t_uv_arr = _to_float_array(t_uv_vals)
            valid    = np.isfinite(t_uv_arr) & np.isfinite(uv_arr)
            if valid.sum() > 1:
                uv_arr = np.interp(t_arr, t_uv_arr[valid], uv_arr[valid],
                                   left=np.nan, right=np.nan)

        # strip NaN-time rows
        valid   = np.isfinite(t_arr)
        t_arr   = t_arr[valid]
        n       = len(t_arr)
        pot_arr = pot_arr[valid] if len(pot_arr) == len(valid) else np.full(n, np.nan)
        uv_arr  = uv_arr[valid]  if len(uv_arr)  == len(valid) else np.full(n, np.nan)

        if t_arr.size == 0:
            continue

        result[sheet_name] = EchemData(
            sheet_name      = sheet_name,
            time            = t_arr,
            potential       = pot_arr,
            uv              = uv_arr,
            potential_label = header[pot_col] if pot_col is not None else "Potential [V]",
            uv_label        = header[uv_col]  if uv_col  is not None else "UV/VIS",
        )

    wb.close()
    return result


def best_sheet_match(sheet_names: list[str], csv_filename: str) -> str:
    """Return the sheet name most similar to the CSV filename stem."""
    stem   = os.path.splitext(csv_filename)[0]
    scored = sorted(sheet_names,
                    key=lambda s: _similarity(s, stem),
                    reverse=True)
    return scored[0] if scored else sheet_names[0]


# ─────────────────────────────────────────────────────────────────────────────
# Data descriptors
# ─────────────────────────────────────────────────────────────────────────────

class Slice:
    _id_counter = 0

    def __init__(self, kind: str, index: int, color: str, label: str = ""):
        Slice._id_counter += 1
        self.uid   = Slice._id_counter
        self.kind  = kind          # "vertical" | "horizontal"
        self.index = index
        self.color = color
        self.label = label or f"{'W' if kind == 'vertical' else 'T'}{self.uid}"
        self.line_obj: Optional[Line2D] = None
        self.visible = True


class Region:
    _id_counter = 0

    def __init__(self, ti0: int, ti1: int, wi0: int, wi1: int,
                 color: str, label: str = ""):
        Region._id_counter += 1
        self.uid = Region._id_counter
        self.ti0, self.ti1 = ti0, ti1
        self.wi0, self.wi1 = wi0, wi1
        self.color = color
        self.label = label or f"R{self.uid}"
        self.rect_obj:     Optional[Rectangle]        = None
        self.visible       = True
        self.detail_window: Optional[RegionPlotWindow] = None


# ─────────────────────────────────────────────────────────────────────────────
# PlotPanel  –  incremental-update plot abstraction
# ─────────────────────────────────────────────────────────────────────────────

class PlotPanel:
    """
    One Figure/Axes embedded in a Tk frame.

    uid → Line2D cache for data lines (incremental set_data updates).
    Separate uid → Line2D cache for vertical marker lines (axvline).
    """

    def __init__(self, parent_frame: tk.Widget,
                 xlabel: str, ylabel: str, title: str,
                 figsize: tuple[float, float] = (4, 3)):
        self.fig, self.ax = plt.subplots(figsize=figsize, facecolor=T.BG_FIGURE)
        _style_axes(self.ax, xlabel, ylabel, title)

        self.canvas = FigureCanvasTkAgg(self.fig, master=parent_frame)
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        self._lines:         dict[int, Line2D] = {}
        self._markers:       dict[int, Line2D] = {}
        self._legend_labels: dict[int, str]    = {}

    # ── data lines ────────────────────────────────────────────────────────

    def upsert(self, uid: int,
               xdata: np.ndarray, ydata: np.ndarray,
               label: str = "", color: str = "#888",
               lw: float = 1.4, ls: str = "-",
               visible: bool = True) -> Line2D:
        if uid in self._lines:
            line = self._lines[uid]
            line.set_data(xdata, ydata)
            line.set_color(color); line.set_linewidth(lw)
            line.set_linestyle(ls); line.set_visible(visible)
            if label:
                line.set_label(label)
        else:
            (line,) = self.ax.plot(xdata, ydata, color=color, lw=lw,
                                   ls=ls, label=label, visible=visible)
            self._lines[uid] = line
        self._legend_labels[uid] = label
        self._rescale()
        return line

    def remove(self, uid: int):
        for store in (self._lines, self._markers):
            ln = store.pop(uid, None)
            if ln is not None:
                try:
                    ln.remove()
                except Exception:
                    pass
        self._legend_labels.pop(uid, None)
        self._rebuild_legend()
        self._rescale()
        self.canvas.draw_idle()

    def clear(self):
        for uid in list(self._lines):
            self.remove(uid)

    def redraw(self):
        self._rebuild_legend()
        self.canvas.draw_idle()

    def set_labels(self, xlabel: str, ylabel: str, title: str):
        _style_axes(self.ax, xlabel, ylabel, title)

    def bind_motion(self, callback):
        self.fig.canvas.mpl_connect("motion_notify_event", callback)

    # ── vertical marker lines ─────────────────────────────────────────────

    def upsert_marker(self, uid: int, xval: float,
                      color: str = "#888", lw: float = 1.2,
                      ls: str = "--", alpha: float = 0.75,
                      visible: bool = True):
        if uid in self._markers:
            m = self._markers[uid]
            m.set_xdata([xval, xval])
            m.set_color(color); m.set_linewidth(lw)
            m.set_linestyle(ls); m.set_alpha(alpha)
            m.set_visible(visible)
        else:
            m = self.ax.axvline(xval, color=color, lw=lw, ls=ls,
                                alpha=alpha, visible=visible)
            self._markers[uid] = m

    def remove_marker(self, uid: int):
        m = self._markers.pop(uid, None)
        if m is not None:
            try:
                m.remove()
            except Exception:
                pass
        self.canvas.draw_idle()

    def clear_markers(self):
        for uid in list(self._markers):
            self.remove_marker(uid)

    # ── internal ──────────────────────────────────────────────────────────

    def _rescale(self):
        self.ax.relim()
        self.ax.autoscale_view()

    def _rebuild_legend(self):
        visible = [l for l in self._lines.values()
                   if l.get_visible() and l.get_label()]
        if visible:
            self.ax.legend(handles=visible,
                           fontsize=T.FONT_SIZE_SMALL,
                           facecolor=T.BG_PANEL,
                           labelcolor=T.FG_LABEL,
                           edgecolor=T.COLOR_SPINE)
        else:
            leg = self.ax.get_legend()
            if leg:
                leg.remove()


# ─────────────────────────────────────────────────────────────────────────────
# EchemPanel  –  UV + potential plots below the heatmap
# ─────────────────────────────────────────────────────────────────────────────

_ECHEM_LINE_UID = -1   # stable uid for the single echem data line


class EchemPanel:
    """
    Two vertically-stacked PlotPanels (UV-vis intensity, potential).

    load(data)  – replace displayed data.
    sync_markers(slices, regions, raman_times)  – mirror time positions.
    """

    def __init__(self, parent_frame: tk.Widget):
        paned = tk.PanedWindow(parent_frame, orient=tk.VERTICAL,
                               bg=T.COLOR_SASH, sashwidth=4)
        paned.pack(fill=tk.BOTH, expand=True)

        uv_frame = tk.Frame(paned, bg=T.BG_APP)
        paned.add(uv_frame, minsize=100)
        self.uv_panel = PlotPanel(uv_frame,
                                  xlabel="Time (s)", ylabel="UV/VIS",
                                  title="UV/VIS intensity vs time",
                                  figsize=(6, 2))

        pot_frame = tk.Frame(paned, bg=T.BG_APP)
        paned.add(pot_frame, minsize=100)
        self.pot_panel = PlotPanel(pot_frame,
                                   xlabel="Time (s)", ylabel="Potential [V]",
                                   title="Potential vs time",
                                   figsize=(6, 2))

    # ── public ────────────────────────────────────────────────────────────

    def load(self, data: EchemData):
        uv_color  = T.SLICE_COLORS[2] if len(T.SLICE_COLORS) > 2 else "#7c3aed"
        pot_color = T.SLICE_COLORS[3] if len(T.SLICE_COLORS) > 3 else "#b45309"

        self.uv_panel.upsert(_ECHEM_LINE_UID,
                             data.time, data.uv,
                             label=data.uv_label,
                             color=uv_color, lw=T.LINE_WIDTH, ls=T.LINE_STYLE)
        self.uv_panel.set_labels("Time (s)", data.uv_label,
                                 f"{data.sheet_name}  –  {data.uv_label}")
        self.uv_panel.redraw()

        self.pot_panel.upsert(_ECHEM_LINE_UID,
                              data.time, data.potential,
                              label=data.potential_label,
                              color=pot_color, lw=T.LINE_WIDTH, ls=T.LINE_STYLE)
        self.pot_panel.set_labels("Time (s)", data.potential_label,
                                  f"{data.sheet_name}  –  {data.potential_label}")
        self.pot_panel.redraw()

    def sync_markers(self, slices: list[Slice],
                     regions: list[Region],
                     raman_times: np.ndarray):
        """Mirror Raman slice/region time positions as axvline markers."""
        for panel in (self.uv_panel, self.pot_panel):
            panel.clear_markers()

        if raman_times.size == 0:
            return

        for s in slices:
            if s.kind != "vertical" or not s.visible:
                continue
            t_val = raman_times[s.index]
            for panel in (self.uv_panel, self.pot_panel):
                panel.upsert_marker(s.uid, t_val,
                                    color=s.color, lw=1.2, ls="--", alpha=0.75)
                panel.canvas.draw_idle()

        for rg in regions:
            if not rg.visible:
                continue
            for marker_uid, t_idx, ls in [
                (rg.uid * 10000,     rg.ti0, "-."),
                (rg.uid * 10000 + 1, rg.ti1, "-."),
            ]:
                t_val = raman_times[t_idx]
                for panel in (self.uv_panel, self.pot_panel):
                    panel.upsert_marker(marker_uid, t_val,
                                        color=rg.color, lw=1.4, ls=ls, alpha=0.65)
                    panel.canvas.draw_idle()


# ─────────────────────────────────────────────────────────────────────────────
# RegionPlotWindow  –  per-region detail window
# ─────────────────────────────────────────────────────────────────────────────

class RegionPlotWindow:
    def __init__(self, parent: tk.Tk, region: Region,
                 wavenumbers: np.ndarray, times: np.ndarray,
                 intensity: np.ndarray):
        self._region = region
        self._wn, self._t, self._Z = wavenumbers, times, intensity

        top = tk.Toplevel(parent)
        top.title(f"Region {region.label} – detail")
        top.geometry("600x540")
        top.configure(bg=T.BG_APP)
        self._top = top
        top.protocol("WM_DELETE_WINDOW", self._on_close)

        paned = tk.PanedWindow(top, orient=tk.VERTICAL,
                               bg=T.COLOR_SASH, sashwidth=5)
        paned.pack(fill=tk.BOTH, expand=True)

        spec_f = tk.Frame(paned, bg=T.BG_APP)
        paned.add(spec_f, minsize=200)
        self._spec_panel = PlotPanel(spec_f,
                                     xlabel="Wavenumber (cm⁻¹)", ylabel="Intensity",
                                     title=f"[{region.label}] Mean spectrum")

        time_f = tk.Frame(paned, bg=T.BG_APP)
        paned.add(time_f, minsize=200)
        self._time_panel = PlotPanel(time_f,
                                     xlabel="Time (s)", ylabel="Intensity",
                                     title=f"[{region.label}] Mean time trace")
        self.refresh()

    def refresh(self):
        rg, wn, t, Z = self._region, self._wn, self._t, self._Z
        if Z.size == 0:
            return
        mean_spec  = np.nanmean(Z[rg.wi0:rg.wi1 + 1, rg.ti0:rg.ti1 + 1], axis=1)
        t_lo, t_hi = t[rg.ti0], t[rg.ti1]
        self._spec_panel.upsert(0, wn[rg.wi0:rg.wi1 + 1], mean_spec,
                                label=f"t:[{t_lo:.3g},{t_hi:.3g}]",
                                color=rg.color, lw=T.LINE_WIDTH, ls=T.LINE_STYLE)
        self._spec_panel.set_labels("Wavenumber (cm⁻¹)", "Intensity",
                                    f"[{rg.label}] Mean spectrum  t:[{t_lo:.3g},{t_hi:.3g}]")
        self._spec_panel.redraw()

        mean_trace = np.nanmean(Z[rg.wi0:rg.wi1 + 1, rg.ti0:rg.ti1 + 1], axis=0)
        w_lo = wn[min(rg.wi0, rg.wi1)];  w_hi = wn[max(rg.wi0, rg.wi1)]
        self._time_panel.upsert(0, t[rg.ti0:rg.ti1 + 1], mean_trace,
                                label=f"wn:[{w_lo:.1f},{w_hi:.1f}]",
                                color=rg.color, lw=T.LINE_WIDTH, ls=T.LINE_STYLE)
        self._time_panel.set_labels("Time (s)", "Intensity",
                                    f"[{rg.label}] Mean time trace  wn:[{w_lo:.1f},{w_hi:.1f}]")
        self._time_panel.redraw()

    def update_color(self):
        for panel in (self._spec_panel, self._time_panel):
            ln = panel._lines.get(0)
            if ln:
                ln.set_color(self._region.color)
                panel.redraw()

    def destroy(self):
        try:
            self._top.destroy()
        except Exception:
            pass

    def _on_close(self):
        self._region.detail_window = None
        self.destroy()


# ─────────────────────────────────────────────────────────────────────────────
# AnalysisWindow
# ─────────────────────────────────────────────────────────────────────────────

class AnalysisWindow:

    MODE_SLICE  = "slice"
    MODE_REGION = "region"

    def __init__(self, root: tk.Tk, path: str):
        self.root  = root
        self.path  = path
        self.fname = os.path.basename(path)

        self.wavenumbers: np.ndarray = np.array([])
        self.times:       np.ndarray = np.array([])
        self.intensity:   np.ndarray = np.array([])
        self.max_intensity:    float = 1.0

        self.slices:         list[Slice]        = []
        self._drag_slice:    Optional[Slice]    = None
        self._focused_slice: Optional[Slice]    = None
        self._slice_widgets: dict[int, dict]    = {}

        self.regions:          list[Region]       = []
        self._focused_region:  Optional[Region]   = None
        self._region_widgets:  dict[int, dict]    = {}

        self._mode             = self.MODE_SLICE
        self._draw_start:      Optional[tuple[float, float]] = None
        self._draw_rect_patch: Optional[Rectangle]           = None

        self._cmap        = T.DEFAULT_CMAP
        self._cbar        = None
        self._heatmap_img = None

        self._echem_sheets: dict[str, EchemData] = {}
        self._echem_path:   Optional[str]         = None

        # ── blit / drag performance state ─────────────────────────────────
        self._heat_bg:            Optional[object] = None
        self._blit_artists:       list              = []
        self._last_side_redraw_t: float             = 0.0

        self._build_ui()
        self._load_data()

    # ─────────────────────────────────────────────────────────────────────
    # UI construction
    # ─────────────────────────────────────────────────────────────────────

    def _build_ui(self):
        r = self.root
        r.title(f"Raman Analysis – {self.fname}")
        r.geometry("1280x980")
        r.minsize(900, 700)
        r.configure(bg=T.BG_APP)

        # ── top toolbar ──────────────────────────────────────────────────
        toolbar = tk.Frame(r, bg=T.BG_TOOLBAR, pady=5)
        toolbar.pack(fill=tk.X, side=tk.TOP)

        tk.Label(toolbar, text=f"  {self.fname}",
                 bg=T.BG_TOOLBAR, fg=T.FG_TITLE,
                 font=(T.FONT_MONO, T.FONT_SIZE_BODY, "bold")).pack(side=tk.LEFT)

        tk.Label(toolbar, text="  cmap:",
                 bg=T.BG_TOOLBAR, fg=T.FG_SUBTLE,
                 font=(T.FONT_MONO, T.FONT_SIZE_SMALL)).pack(side=tk.LEFT)

        self._cmap_var = tk.StringVar(value=self._cmap)
        cmap_menu = ttk.Combobox(toolbar, textvariable=self._cmap_var,
                                 values=T.AVAILABLE_CMAPS, width=10, state="readonly")
        cmap_menu.pack(side=tk.LEFT, padx=(2, 12))
        cmap_menu.bind("<<ComboboxSelected>>", lambda e: self._refresh_heatmap())

        btn_cfg = dict(font=(T.FONT_MONO, T.FONT_SIZE_SMALL),
                       relief=tk.FLAT, padx=8, bd=0)

        tk.Button(toolbar, text="+ Spectrum slice (V)",
                  bg="#dbeafe", fg="#1e40af",
                  activebackground="#bfdbfe", activeforeground="#1e3a8a",
                  command=self._add_vertical_slice, **btn_cfg).pack(side=tk.LEFT, padx=2)

        tk.Button(toolbar, text="+ Time trace (H)",
                  bg="#dcfce7", fg="#166534",
                  activebackground="#bbf7d0", activeforeground="#14532d",
                  command=self._add_horizontal_slice, **btn_cfg).pack(side=tk.LEFT, padx=2)

        tk.Label(toolbar, text="  |  ", bg=T.BG_TOOLBAR,
                 fg=T.COLOR_SPINE).pack(side=tk.LEFT)

        self._region_btn_text = tk.StringVar(value="[ ] Draw region")
        self._region_btn = tk.Button(
            toolbar, textvariable=self._region_btn_text,
            bg="#fef9c3", fg="#854d0e",
            activebackground="#fef08a", activeforeground="#713f12",
            command=self._toggle_region_mode, **btn_cfg)
        self._region_btn.pack(side=tk.LEFT, padx=2)

        tk.Button(toolbar, text="x Remove selected",
                  bg="#fee2e2", fg="#991b1b",
                  activebackground="#fecaca", activeforeground="#7f1d1d",
                  command=self._remove_focused, **btn_cfg).pack(side=tk.LEFT, padx=2)

        self._info_var = tk.StringVar(value="")
        tk.Label(toolbar, textvariable=self._info_var,
                 bg=T.BG_TOOLBAR, fg=T.FG_SUBTLE,
                 font=(T.FONT_MONO, T.FONT_SIZE_SMALL)).pack(side=tk.RIGHT, padx=12)

        # ── echem toolbar ─────────────────────────────────────────────────
        echem_bar = tk.Frame(r, bg="#f0fdf4", pady=4)
        echem_bar.pack(fill=tk.X, side=tk.TOP)

        tk.Label(echem_bar, text="  Echem xlsx:",
                 bg="#f0fdf4", fg=T.FG_SUBTLE,
                 font=(T.FONT_MONO, T.FONT_SIZE_SMALL)).pack(side=tk.LEFT)

        self._echem_path_var = tk.StringVar(value="(none loaded)")
        tk.Label(echem_bar, textvariable=self._echem_path_var,
                 bg="#f0fdf4", fg="#166534",
                 font=(T.FONT_MONO, T.FONT_SIZE_SMALL)).pack(side=tk.LEFT, padx=(4, 12))

        tk.Button(echem_bar, text="Load xlsx…",
                  bg="#bbf7d0", fg="#14532d",
                  activebackground="#86efac", activeforeground="#14532d",
                  command=self._browse_echem, **btn_cfg).pack(side=tk.LEFT, padx=2)

        tk.Label(echem_bar, text="  Sheet:",
                 bg="#f0fdf4", fg=T.FG_SUBTLE,
                 font=(T.FONT_MONO, T.FONT_SIZE_SMALL)).pack(side=tk.LEFT, padx=(10, 2))

        self._sheet_var   = tk.StringVar(value="")
        self._sheet_combo = ttk.Combobox(echem_bar, textvariable=self._sheet_var,
                                         values=[], width=24, state="readonly")
        self._sheet_combo.pack(side=tk.LEFT, padx=(0, 6))
        self._sheet_combo.bind("<<ComboboxSelected>>", self._on_sheet_selected)

        self._echem_info_var = tk.StringVar(value="")
        tk.Label(echem_bar, textvariable=self._echem_info_var,
                 bg="#f0fdf4", fg=T.FG_SUBTLE,
                 font=(T.FONT_MONO, T.FONT_SIZE_SMALL)).pack(side=tk.LEFT, padx=6)

        # ── bottom bar ────────────────────────────────────────────────────
        # IMPORTANT: pack the bottom bar BEFORE the expanding PanedWindow.
        # Tk's pack geometry manager allocates space to side=BOTTOM widgets
        # first (in reverse pack order for BOTTOM), so by packing this before
        # the main paned window we guarantee it always gets its natural height
        # and is never squeezed out when the window is small.
        bottom = tk.Frame(r, bg=T.BG_SLICE_BAR)
        bottom.pack(fill=tk.X, side=tk.BOTTOM)

        sr = tk.Frame(bottom, bg=T.BG_SLICE_BAR, pady=3)
        sr.pack(fill=tk.X)
        tk.Label(sr, text=" Slices: ",
                 bg=T.BG_SLICE_BAR, fg=T.FG_SUBTLE,
                 font=(T.FONT_MONO, T.FONT_SIZE_SMALL)).pack(side=tk.LEFT)
        self._slice_list_frame = tk.Frame(sr, bg=T.BG_SLICE_BAR)
        self._slice_list_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)

        _RRB = "#fefce8"
        rr = tk.Frame(bottom, bg=_RRB, pady=3)
        rr.pack(fill=tk.X)
        tk.Label(rr, text=" Regions:",
                 bg=_RRB, fg=T.FG_SUBTLE,
                 font=(T.FONT_MONO, T.FONT_SIZE_SMALL)).pack(side=tk.LEFT)
        self._region_list_frame = tk.Frame(rr, bg=_RRB)
        self._region_list_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self._REGION_ROW_BG = _RRB

        # ── outer vertical paned: upper (heatmap+side) / lower (echem) ───
        # Packed AFTER bottom so the paned window fills remaining space only.
        outer = tk.PanedWindow(r, orient=tk.VERTICAL,
                               bg=T.COLOR_SASH, sashwidth=6)
        outer.pack(fill=tk.BOTH, expand=True)

        # upper half
        upper = tk.Frame(outer, bg=T.BG_APP)
        outer.add(upper, minsize=360)

        paned = tk.PanedWindow(upper, orient=tk.HORIZONTAL,
                               bg=T.COLOR_SASH, sashwidth=6, sashrelief=tk.FLAT)
        paned.pack(fill=tk.BOTH, expand=True)

        left = tk.PanedWindow(paned, orient=tk.VERTICAL,
                               bg=T.COLOR_SASH, sashwidth=6, sashrelief=tk.FLAT)
        paned.add(left, minsize=400)

        heat_frame = tk.Frame(left, bg=T.BG_APP)
        left.add(heat_frame, minsize=400)

        self._fig_heat, self._ax_heat = plt.subplots(figsize=(6, 5),
                                                      facecolor=T.BG_FIGURE)
        _style_axes(self._ax_heat, "Time (s)", "Wavenumber (cm⁻¹)",
                    "Raman Intensity Heatmap")
        self._canvas_heat = FigureCanvasTkAgg(self._fig_heat, master=heat_frame)
        self._canvas_heat.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        self._fig_heat.canvas.mpl_connect("button_press_event",   self._on_heat_press)
        self._fig_heat.canvas.mpl_connect("motion_notify_event",  self._on_heat_motion)
        self._fig_heat.canvas.mpl_connect("button_release_event", self._on_heat_release)
        self._fig_heat.canvas.mpl_connect("scroll_event",         self._on_heat_scroll)

        # lower half – echem
        echem_outer = tk.Frame(left, bg=T.BG_APP)
        left.add(echem_outer, minsize=220)

        echem_hdr = tk.Frame(echem_outer, bg="#f0fdf4", pady=2)
        echem_hdr.pack(fill=tk.X)
        tk.Label(echem_hdr, text="  Electrochemistry",
                 bg="#f0fdf4", fg="#166534",
                 font=(T.FONT_MONO, T.FONT_SIZE_SMALL, "bold")).pack(side=tk.LEFT)
        self._echem_sheet_lbl = tk.StringVar(value="")
        tk.Label(echem_hdr, textvariable=self._echem_sheet_lbl,
                 bg="#f0fdf4", fg=T.FG_SUBTLE,
                 font=(T.FONT_MONO, T.FONT_SIZE_SMALL)).pack(side=tk.LEFT, padx=6)

        self._echem_panel = EchemPanel(echem_outer)

        right = tk.Frame(paned, bg=T.BG_APP)
        paned.add(right, minsize=340)

        rpaned = tk.PanedWindow(right, orient=tk.VERTICAL,
                                bg=T.COLOR_SASH, sashwidth=5)
        rpaned.pack(fill=tk.BOTH, expand=True)

        spec_f = tk.Frame(rpaned, bg=T.BG_APP)
        rpaned.add(spec_f, minsize=170)
        self._spec_panel = PlotPanel(spec_f,
                                     xlabel="Wavenumber (cm⁻¹)", ylabel="Intensity",
                                     title="Spectra (V-slices)")
        self._spec_panel.bind_motion(self._on_spec_motion)

        time_f = tk.Frame(rpaned, bg=T.BG_APP)
        rpaned.add(time_f, minsize=170)
        self._time_panel = PlotPanel(time_f,
                                     xlabel="Time (s)", ylabel="Intensity",
                                     title="Time traces (H-slices)")
        self._time_panel.bind_motion(self._on_time_motion)

        r.bind("<Left>",   lambda e: self._nudge_focused(-1))
        r.bind("<Right>",  lambda e: self._nudge_focused(+1))
        r.bind("<Up>",     lambda e: self._nudge_focused(+1))
        r.bind("<Down>",   lambda e: self._nudge_focused(-1))
        r.bind("<Delete>", lambda e: self._remove_focused())
        r.bind("<Escape>", lambda e: self._escape_pressed())

    # ─────────────────────────────────────────────────────────────────────
    # Raman loading
    # ─────────────────────────────────────────────────────────────────────

    def _load_data(self):
        try:
            wn, t, Z = load_raman_csv(self.path)
        except Exception as exc:
            messagebox.showerror("Load error",
                                 f"Could not read file:\n{self.path}\n\n{exc}",
                                 parent=self.root)
            self.root.destroy()
            traceback.print_exc()
            return
        self.wavenumbers, self.times, self.intensity = wn, t, Z
        self.max_intensity = float(np.amax(self.intensity))
        self._refresh_heatmap()
        self._add_vertical_slice()
        self._add_horizontal_slice()

        # auto-probe same-name xlsx in same directory
        candidate = os.path.splitext(self.path)[0] + ".xlsx"
        if os.path.isfile(candidate):
            self._load_echem_xlsx(candidate)

    # ─────────────────────────────────────────────────────────────────────
    # Echem loading
    # ─────────────────────────────────────────────────────────────────────

    def _browse_echem(self):
        path = filedialog.askopenfilename(
            title="Select electrochemistry xlsx",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
            initialdir=os.path.dirname(self.path),
            parent=self.root,
        )
        if path:
            self._load_echem_xlsx(path)

    def _load_echem_xlsx(self, path: str):
        try:
            sheets = load_echem_xlsx(path)
        except ImportError as exc:
            messagebox.showerror("Missing dependency", str(exc), parent=self.root)
            return
        except Exception as exc:
            messagebox.showerror("Load error",
                                 f"Could not read xlsx:\n{path}\n\n{exc}",
                                 parent=self.root)
            traceback.print_exc()
            return
        if not sheets:
            messagebox.showwarning("No data",
                                   "No usable sheets found in the xlsx file.",
                                   parent=self.root)
            return

        self._echem_sheets = sheets
        self._echem_path   = path
        self._echem_path_var.set(os.path.basename(path))

        names = list(sheets.keys())
        self._sheet_combo.configure(values=names)
        best = best_sheet_match(names, self.fname)
        self._sheet_var.set(best)
        self._apply_sheet(best)

    def _on_sheet_selected(self, _event=None):
        name = self._sheet_var.get()
        if name in self._echem_sheets:
            self._apply_sheet(name)

    def _apply_sheet(self, name: str):
        data = self._echem_sheets.get(name)
        if data is None:
            return
        self._echem_panel.load(data)
        n = len(data.time)
        self._echem_sheet_lbl.set(f"  {name}  ({n} pts)")
        self._echem_info_var.set(
            f"t=[{data.time[0]:.3g}, {data.time[-1]:.3g}] s  "
            f"| {data.potential_label}: [{np.nanmin(data.potential):.3g}, "
            f"{np.nanmax(data.potential):.3g}] V"
            if data.potential.size else ""
        )
        self._sync_echem_markers()

    def _sync_echem_markers(self):
        self._echem_panel.sync_markers(self.slices, self.regions, self.times)

    # ─────────────────────────────────────────────────────────────────────
    # Heatmap rendering
    # ─────────────────────────────────────────────────────────────────────

    def _refresh_heatmap(self):
        """Full redraw of the heatmap (expensive).
        Called only when data/cmap changes, never during drag.
        Invalidates the blit background cache."""
        if self.intensity.size == 0:
            return
        if self._cbar is not None:
            self._cbar.remove()
        self._cmap = self._cmap_var.get()
        ax = self._ax_heat
        ax.cla()
        _style_axes(ax, "Time (s)", "Wavenumber (cm⁻¹)", "Raman Intensity Heatmap")
        t0, t1 = self.times[0], self.times[-1]
        w0, w1 = self.wavenumbers[0], self.wavenumbers[-1]
        self._heatmap_img = ax.imshow(
            self.intensity, aspect="auto",
            origin="lower" if w1 > w0 else "upper",
            extent=(t0, t1, w0, w1),
            cmap=self._cmap, interpolation="none",
            vmax=self.max_intensity * 0.5  # TODO: make factor user accessible
        )
        self._cbar = self._fig_heat.colorbar(self._heatmap_img, ax=ax, location="right")
        for s in self.slices:
            s.line_obj = None
            self._draw_slice_line(s)
        for rg in self.regions:
            rg.rect_obj = None
            self._draw_region_rect(rg)

        # Invalidate blit cache — recaptured on next drag start
        self._heat_bg = None
        # Use draw() not draw_idle() here so the canvas is fully rendered
        # before any subsequent copy_from_bbox call.
        self._canvas_heat.draw()

    def _draw_slice_line(self, s: Slice):
        """Create or update the Line2D for a slice on the heatmap axes."""
        focused = s is self._focused_slice
        lw = 2.2 if focused else 1.4
        ls = "-"  if focused else "--"
        if s.kind == "vertical":
            val = self.times[s.index]
            if s.line_obj is None:
                s.line_obj = self._ax_heat.axvline(val, color=s.color, lw=lw,
                                                   ls=ls, animated=False)
            else:
                s.line_obj.set_xdata([val, val])
        else:
            val = self.wavenumbers[s.index]
            if s.line_obj is None:
                s.line_obj = self._ax_heat.axhline(val, color=s.color, lw=lw,
                                                   ls=ls, animated=False)
            else:
                s.line_obj.set_ydata([val, val])
        s.line_obj.set_color(s.color); s.line_obj.set_linewidth(lw)
        s.line_obj.set_linestyle(ls);  s.line_obj.set_visible(s.visible)

    def _draw_region_rect(self, rg: Region):
        t_lo = self.times[rg.ti0];  t_hi = self.times[rg.ti1]
        w_lo = self.wavenumbers[min(rg.wi0, rg.wi1)]
        w_hi = self.wavenumbers[max(rg.wi0, rg.wi1)]
        focused    = rg is self._focused_region
        lw         = 2.2 if focused else 1.4
        alpha_fill = 0.30 if focused else 0.20
        if rg.rect_obj is None:
            rg.rect_obj = Rectangle(
                (t_lo, w_lo), t_hi - t_lo, w_hi - w_lo,
                linewidth=lw, edgecolor=rg.color,
                facecolor=rg.color, alpha=alpha_fill,
                linestyle="-", zorder=3, animated=False)
            self._ax_heat.add_patch(rg.rect_obj)
        else:
            rg.rect_obj.set_xy((t_lo, w_lo))
            rg.rect_obj.set_width(t_hi - t_lo)
            rg.rect_obj.set_height(w_hi - w_lo)
            rg.rect_obj.set_edgecolor(rg.color)
            rg.rect_obj.set_facecolor(rg.color)
            rg.rect_obj.set_linewidth(lw)
            rg.rect_obj.set_alpha(alpha_fill)
        rg.rect_obj.set_visible(rg.visible)

    # ─────────────────────────────────────────────────────────────────────
    # Blit-based drag helpers
    # ─────────────────────────────────────────────────────────────────────

    def _blit_drag_start(self, s: Slice):
        """
        Begin blit session for slice s.

        Strategy:
          1. Mark the dragged line as animated=True (excluded from normal
             draw cycle so it won't appear in the captured background).
          2. Do a full canvas.draw() so all *other* artists are rendered.
          3. Capture the result via copy_from_bbox — this is our static bg.
          4. Immediately restore and re-draw the line so the user sees it.
        """
        line = s.line_obj
        if line is None:
            return
        line.set_animated(True)
        self._fig_heat.canvas.draw()
        self._heat_bg      = self._fig_heat.canvas.copy_from_bbox(self._ax_heat.bbox)
        self._blit_artists = [line]
        # Paint the line back immediately so there's no flash
        self._fig_heat.canvas.restore_region(self._heat_bg)
        self._ax_heat.draw_artist(line)
        self._fig_heat.canvas.blit(self._ax_heat.bbox)

    def _blit_drag_move(self, s: Slice):
        """
        Fast per-motion update: restore bg bitmap, re-draw only the
        moving line, push just the axes bbox to the screen.
        No Python-side rasterisation of the heatmap image occurs.
        """
        if self._heat_bg is None or not self._blit_artists:
            self._canvas_heat.draw_idle()   # graceful fallback
            return
        canvas = self._fig_heat.canvas
        canvas.restore_region(self._heat_bg)
        for artist in self._blit_artists:
            self._ax_heat.draw_artist(artist)
        canvas.blit(self._ax_heat.bbox)

    def _blit_drag_end(self, s: Slice):
        """
        End blit session: restore line to non-animated, do a proper
        full draw so the final position is permanently composited into
        the canvas (and the background cache is valid again).
        """
        line = s.line_obj
        if line is not None:
            line.set_animated(False)
        self._heat_bg      = None
        self._blit_artists = []
        self._canvas_heat.draw_idle()

    # ─────────────────────────────────────────────────────────────────────
    # PlotPanel update helpers
    # ─────────────────────────────────────────────────────────────────────

    def _upsert_slice_in_panel(self, s: Slice, *, throttled: bool = False):
        """
        Push updated slice data to the spectrum / time side panel.

        throttled=True  →  skip the redraw if fewer than
        SIDE_PANEL_THROTTLE_MS ms have elapsed since the last update.
        This prevents a build-up of queued redraws during fast drags
        while still giving the user a live preview at ~25 fps.
        """
        if throttled:
            now = time.monotonic()
            if (now - self._last_side_redraw_t) * 1000 < SIDE_PANEL_THROTTLE_MS:
                return
            self._last_side_redraw_t = now

        if not s.visible or self.intensity.size == 0:
            panel = self._spec_panel if s.kind == "vertical" else self._time_panel
            if s.uid in panel._lines:
                panel._lines[s.uid].set_visible(False)
                panel.redraw()
            self._sync_echem_markers()
            return

        if s.kind == "vertical":
            self._spec_panel.upsert(
                s.uid, self.wavenumbers, self.intensity[:, s.index],
                label=f"{s.label}  t={self.times[s.index]:.3g}",
                color=s.color, lw=T.LINE_WIDTH, ls=T.LINE_STYLE)
            self._spec_panel.redraw()
        else:
            self._time_panel.upsert(
                s.uid, self.times, self.intensity[s.index, :],
                label=f"{s.label}  wn={self.wavenumbers[s.index]:.1f}",
                color=s.color, lw=T.LINE_WIDTH, ls=T.LINE_STYLE)
            self._time_panel.redraw()
        self._sync_echem_markers()

    def _remove_slice_from_panel(self, s: Slice):
        (self._spec_panel if s.kind == "vertical" else self._time_panel).remove(s.uid)
        self._sync_echem_markers()

    # ─────────────────────────────────────────────────────────────────────
    # Slice management
    # ─────────────────────────────────────────────────────────────────────

    def _used_colors(self) -> list[str]:
        return [s.color for s in self.slices] + [rg.color for rg in self.regions]

    def _add_vertical_slice(self):
        idx = len(self.times) // 2 if self.times.size else 0
        s   = Slice("vertical", idx, _next_color(self._used_colors()))
        self.slices.append(s)
        self._draw_slice_line(s)
        self._focused_slice = s; self._focused_region = None
        self._rebuild_slice_list()
        self._upsert_slice_in_panel(s)

    def _add_horizontal_slice(self):
        idx = len(self.wavenumbers) // 2 if self.wavenumbers.size else 0
        s   = Slice("horizontal", idx, _next_color(self._used_colors()))
        self.slices.append(s)
        self._draw_slice_line(s)
        self._focused_slice = s; self._focused_region = None
        self._rebuild_slice_list()
        self._upsert_slice_in_panel(s)

    def _set_focused_slice(self, s: Optional[Slice]):
        prev_sl = self._focused_slice
        prev_rg = self._focused_region
        self._focused_slice = s; self._focused_region = None
        if prev_rg is not None:
            self._draw_region_rect(prev_rg); self._update_region_widgets(prev_rg)
        for sl in [prev_sl, s]:
            if sl is not None:
                self._draw_slice_line(sl); self._update_slice_widgets(sl)
        self._canvas_heat.draw_idle()

    def _nudge_focused(self, delta: int):
        s = self._focused_slice
        if s is None:
            return
        arr = self.times if s.kind == "vertical" else self.wavenumbers
        s.index = int(np.clip(s.index + delta, 0, len(arr) - 1))
        self._draw_slice_line(s)
        self._canvas_heat.draw_idle()
        self._upsert_slice_in_panel(s)
        self._update_slice_widgets(s)

    def _remove_focused(self):
        if self._focused_slice is not None:
            self._remove_slice(self._focused_slice)
        elif self._focused_region is not None:
            self._remove_region(self._focused_region)

    def _remove_slice(self, s: Slice):
        if s.line_obj is not None:
            try:
                s.line_obj.remove()
            except Exception:
                pass
        self.slices.remove(s)
        if self._focused_slice is s:
            self._focused_slice = None
        self._canvas_heat.draw_idle()
        self._rebuild_slice_list()
        self._remove_slice_from_panel(s)

    def _escape_pressed(self):
        if self._mode == self.MODE_REGION and self._draw_start is not None:
            self._cancel_draw()
        elif self._focused_slice is not None:
            self._set_focused_slice(None)
        elif self._focused_region is not None:
            self._set_focused_region(None)

    # ─────────────────────────────────────────────────────────────────────
    # Region management
    # ─────────────────────────────────────────────────────────────────────

    def _toggle_region_mode(self):
        if self._mode == self.MODE_SLICE:
            self._mode = self.MODE_REGION
            self._region_btn.configure(bg="#fde047", fg="#713f12")
            self._region_btn_text.set("[ ] Drawing… (click+drag)")
            self._canvas_heat.get_tk_widget().configure(cursor="crosshair")
        else:
            self._cancel_draw()
            self._mode = self.MODE_SLICE
            self._region_btn.configure(bg="#fef9c3", fg="#854d0e")
            self._region_btn_text.set("[ ] Draw region")
            self._canvas_heat.get_tk_widget().configure(cursor="")

    def _cancel_draw(self):
        self._draw_start = None
        if self._draw_rect_patch is not None:
            try:
                self._draw_rect_patch.remove()
            except Exception:
                pass
            self._draw_rect_patch = None
        self._canvas_heat.draw_idle()

    def _commit_region(self, t0_d: float, t1_d: float,
                       w0_d: float, w1_d: float):
        ti0 = int(np.clip(np.searchsorted(self.times, min(t0_d, t1_d)),
                          0, len(self.times) - 1))
        ti1 = int(np.clip(np.searchsorted(self.times, max(t0_d, t1_d)),
                          0, len(self.times) - 1))
        wi0 = int(np.clip(np.searchsorted(self.wavenumbers, min(w0_d, w1_d)),
                          0, len(self.wavenumbers) - 1))
        wi1 = int(np.clip(np.searchsorted(self.wavenumbers, max(w0_d, w1_d)),
                          0, len(self.wavenumbers) - 1))
        if ti0 == ti1 or wi0 == wi1:
            return
        rg = Region(ti0, ti1, wi0, wi1, _next_color(self._used_colors()))
        self.regions.append(rg)
        self._draw_region_rect(rg)
        self._focused_region = rg; self._focused_slice = None
        self._rebuild_region_list()
        self._open_region_detail(rg)
        self._sync_echem_markers()
        self._canvas_heat.draw_idle()

    def _set_focused_region(self, rg: Optional[Region]):
        prev_rg = self._focused_region; prev_sl = self._focused_slice
        self._focused_region = rg; self._focused_slice = None
        if prev_sl is not None:
            self._draw_slice_line(prev_sl); self._update_slice_widgets(prev_sl)
        for r in [prev_rg, rg]:
            if r is not None:
                self._draw_region_rect(r); self._update_region_widgets(r)
        self._canvas_heat.draw_idle()

    def _remove_region(self, rg: Region):
        if rg.rect_obj is not None:
            try:
                rg.rect_obj.remove()
            except Exception:
                pass
        if rg.detail_window is not None:
            rg.detail_window.destroy(); rg.detail_window = None
        self.regions.remove(rg)
        if self._focused_region is rg:
            self._focused_region = None
        self._canvas_heat.draw_idle()
        self._rebuild_region_list()
        self._sync_echem_markers()

    def _open_region_detail(self, rg: Region):
        if rg.detail_window is not None:
            try:
                rg.detail_window._top.lift(); return
            except Exception:
                rg.detail_window = None
        rg.detail_window = RegionPlotWindow(
            self.root, rg, self.wavenumbers, self.times, self.intensity)

    # ─────────────────────────────────────────────────────────────────────
    # Heatmap mouse events
    # ─────────────────────────────────────────────────────────────────────

    def _nearest_slice(self, event) -> Optional[Slice]:
        if event.xdata is None or event.ydata is None:
            return None
        ax = self._ax_heat
        best, best_dist = None, np.inf
        for s in self.slices:
            if s.kind == "vertical":
                d = ax.transData.transform((self.times[s.index], event.ydata))
                c = ax.transData.transform((event.xdata, event.ydata))
                dist = abs(d[0] - c[0])
            else:
                d = ax.transData.transform((event.xdata, self.wavenumbers[s.index]))
                c = ax.transData.transform((event.xdata, event.ydata))
                dist = abs(d[1] - c[1])
            if dist < best_dist:
                best_dist = dist; best = s
        return best if best_dist < 12 else None

    def _on_heat_press(self, event):
        if event.inaxes != self._ax_heat or event.xdata is None:
            return
        if self._mode == self.MODE_REGION:
            self._draw_start      = (event.xdata, event.ydata)
            self._draw_rect_patch = Rectangle(
                (event.xdata, event.ydata), 0, 0,
                linewidth=1.5, edgecolor="#854d0e",
                facecolor="#fde047", alpha=0.25,
                linestyle="--", zorder=4)
            self._ax_heat.add_patch(self._draw_rect_patch)
            return
        s = self._nearest_slice(event)
        if s is not None:
            self._drag_slice = s
            self._set_focused_slice(s)
            # Start blit session — captures heatmap bg without the moving line
            self._blit_drag_start(s)
            return
        for rg in self.regions:
            t_lo = self.times[rg.ti0];  t_hi = self.times[rg.ti1]
            w_lo = self.wavenumbers[min(rg.wi0, rg.wi1)]
            w_hi = self.wavenumbers[max(rg.wi0, rg.wi1)]
            if t_lo <= event.xdata <= t_hi and w_lo <= event.ydata <= w_hi:
                self._set_focused_region(rg); return
        self._set_focused_slice(None)

    def _on_heat_motion(self, event):
        if event.inaxes != self._ax_heat or event.xdata is None:
            return
        ti  = int(np.clip(np.searchsorted(self.times, event.xdata),
                          0, len(self.times) - 1))
        wi  = int(np.clip(np.searchsorted(self.wavenumbers, event.ydata),
                          0, len(self.wavenumbers) - 1))
        val = self.intensity[wi, ti] if self.intensity.size else 0
        self._info_var.set(
            f"t={self.times[ti]:.3g}  wn={self.wavenumbers[wi]:.1f}  I={val:.4g}")

        if self._mode == self.MODE_REGION and self._draw_start is not None:
            x0, y0 = self._draw_start
            if self._draw_rect_patch is not None:
                self._draw_rect_patch.set_xy((min(x0, event.xdata), min(y0, event.ydata)))
                self._draw_rect_patch.set_width(abs(event.xdata - x0))
                self._draw_rect_patch.set_height(abs(event.ydata - y0))
                self._canvas_heat.draw_idle()
            return

        if self._drag_slice is not None:
            s = self._drag_slice
            if s.kind == "vertical":
                idx = int(np.clip(np.searchsorted(self.times, event.xdata),
                                  0, len(self.times) - 1))
            else:
                idx = int(np.clip(np.searchsorted(self.wavenumbers, event.ydata),
                                  0, len(self.wavenumbers) - 1))
            if idx != s.index:
                s.index = idx
                # Move the artist's position directly — no redraw of heatmap
                if s.kind == "vertical":
                    s.line_obj.set_xdata([self.times[idx], self.times[idx]])
                else:
                    s.line_obj.set_ydata([self.wavenumbers[idx], self.wavenumbers[idx]])

                # Fast blit update on the heatmap canvas
                self._blit_drag_move(s)

                # Throttled side-panel update (~25 fps during drag)
                self._upsert_slice_in_panel(s, throttled=True)
                self._update_slice_widgets(s)

    def _on_heat_release(self, event):
        if self._mode == self.MODE_REGION and self._draw_start is not None:
            if event.xdata is not None and event.inaxes == self._ax_heat:
                x0, y0 = self._draw_start
                self._cancel_draw()
                self._commit_region(x0, event.xdata, y0, event.ydata)
            else:
                self._cancel_draw()
            self._mode = self.MODE_SLICE
            self._region_btn.configure(bg="#fef9c3", fg="#854d0e")
            self._region_btn_text.set("[ ] Draw region")
            self._canvas_heat.get_tk_widget().configure(cursor="")
            return

        if self._drag_slice is not None:
            s = self._drag_slice
            self._drag_slice = None
            # End blit mode → full redraw at final position + full panel sync
            self._blit_drag_end(s)
            self._upsert_slice_in_panel(s, throttled=False)
            self._sync_echem_markers()

    def _on_heat_scroll(self, event):
        self._nudge_focused(1 if event.button == "up" else -1)

    # ─────────────────────────────────────────────────────────────────────
    # Side panel motion info
    # ─────────────────────────────────────────────────────────────────────

    def _on_spec_motion(self, event):
        if event.inaxes != self._spec_panel.ax or event.xdata is None:
            return
        wi = int(np.clip(np.searchsorted(self.wavenumbers, event.xdata),
                         0, len(self.wavenumbers) - 1))
        if self.intensity.size:
            row = self.intensity[wi, :]
            self._info_var.set(
                f"wn={self.wavenumbers[wi]:.1f}  "
                f"I_min={np.nanmin(row):.4g}  I_max={np.nanmax(row):.4g}")

    def _on_time_motion(self, event):
        if event.inaxes != self._time_panel.ax or event.xdata is None:
            return
        ti = int(np.clip(np.searchsorted(self.times, event.xdata),
                         0, len(self.times) - 1))
        if self.intensity.size:
            col = self.intensity[:, ti]
            self._info_var.set(
                f"t={self.times[ti]:.3g}  "
                f"I_min={np.nanmin(col):.4g}  I_max={np.nanmax(col):.4g}")

    # ─────────────────────────────────────────────────────────────────────
    # Slice list bar
    # ─────────────────────────────────────────────────────────────────────

    def _rebuild_slice_list(self):
        for w in self._slice_list_frame.winfo_children():
            w.destroy()
        self._slice_widgets.clear()
        for s in self.slices:
            self._slice_widgets[s.uid] = self._create_slice_widgets(s)

    def _create_slice_widgets(self, s: Slice) -> dict:
        focused = s is self._focused_slice
        frame = tk.Frame(self._slice_list_frame, bg=T.BG_SLICE_BAR, padx=3, pady=1)
        frame.pack(side=tk.LEFT)
        inner = tk.Frame(frame, bg=s.color if focused else T.COLOR_SPINE, padx=1, pady=1)
        inner.pack()
        row = tk.Frame(inner, bg=T.BG_PANEL)
        row.pack()

        swatch = tk.Label(row, bg=s.color, width=2, cursor="hand2")
        swatch.pack(side=tk.LEFT)
        swatch.bind("<Button-1>", lambda e, sl=s: self._pick_slice_color(sl))

        arr  = self.times if s.kind == "vertical" else self.wavenumbers
        val  = arr[s.index] if arr.size else 0
        icon = "|" if s.kind == "vertical" else "-"
        lbl  = tk.Label(row, text=f" {icon}{s.label}={val:.3g} ",
                        bg=T.BG_PANEL, fg=s.color,
                        font=(T.FONT_MONO, T.FONT_SIZE_SMALL), cursor="hand2")
        lbl.pack(side=tk.LEFT)
        lbl.bind("<Button-1>", lambda e, sl=s: self._set_focused_slice(sl))

        vis = tk.Label(row, text="O" if s.visible else "o",
                       bg=T.BG_PANEL, fg=T.FG_SUBTLE,
                       font=(T.FONT_MONO, T.FONT_SIZE_SMALL), cursor="hand2")
        vis.pack(side=tk.LEFT)
        vis.bind("<Button-1>", lambda e, sl=s: self._toggle_slice_visible(sl))

        return dict(inner=inner, swatch=swatch, lbl=lbl, vis_btn=vis)

    def _update_slice_widgets(self, s: Slice):
        w = self._slice_widgets.get(s.uid)
        if w is None:
            return
        focused = s is self._focused_slice
        w["inner"].configure(bg=s.color if focused else T.COLOR_SPINE)
        w["swatch"].configure(bg=s.color)
        arr  = self.times if s.kind == "vertical" else self.wavenumbers
        val  = arr[s.index] if arr.size else 0
        icon = "|" if s.kind == "vertical" else "-"
        w["lbl"].configure(text=f" {icon}{s.label}={val:.3g} ", fg=s.color)
        w["vis_btn"].configure(text="O" if s.visible else "o")

    def _pick_slice_color(self, s: Slice):
        c = colorchooser.askcolor(color=s.color, title=f"Color for {s.label}",
                                  parent=self.root)
        if c and c[1]:
            s.color = c[1]
            self._draw_slice_line(s)
            self._canvas_heat.draw_idle()
            self._upsert_slice_in_panel(s)
            self._update_slice_widgets(s)

    def _toggle_slice_visible(self, s: Slice):
        s.visible = not s.visible
        self._draw_slice_line(s)
        self._canvas_heat.draw_idle()
        self._upsert_slice_in_panel(s)
        self._update_slice_widgets(s)

    # ─────────────────────────────────────────────────────────────────────
    # Region list bar
    # ─────────────────────────────────────────────────────────────────────

    def _rebuild_region_list(self):
        for w in self._region_list_frame.winfo_children():
            w.destroy()
        self._region_widgets.clear()
        for rg in self.regions:
            self._region_widgets[rg.uid] = self._create_region_widgets(rg)

    def _create_region_widgets(self, rg: Region) -> dict:
        focused = rg is self._focused_region
        bg = self._REGION_ROW_BG
        frame = tk.Frame(self._region_list_frame, bg=bg, padx=3, pady=1)
        frame.pack(side=tk.LEFT)
        inner = tk.Frame(frame, bg=rg.color if focused else T.COLOR_SPINE, padx=1, pady=1)
        inner.pack()
        row = tk.Frame(inner, bg=T.BG_PANEL)
        row.pack()

        swatch = tk.Label(row, bg=rg.color, width=2, cursor="hand2")
        swatch.pack(side=tk.LEFT)
        swatch.bind("<Button-1>", lambda e, r=rg: self._pick_region_color(r))

        lbl = tk.Label(row, text=self._region_label_text(rg),
                       bg=T.BG_PANEL, fg=rg.color,
                       font=(T.FONT_MONO, T.FONT_SIZE_SMALL), cursor="hand2")
        lbl.pack(side=tk.LEFT)
        lbl.bind("<Button-1>", lambda e, r=rg: self._set_focused_region(r))

        vis = tk.Label(row, text="O" if rg.visible else "o",
                       bg=T.BG_PANEL, fg=T.FG_SUBTLE,
                       font=(T.FONT_MONO, T.FONT_SIZE_SMALL), cursor="hand2")
        vis.pack(side=tk.LEFT)
        vis.bind("<Button-1>", lambda e, r=rg: self._toggle_region_visible(r))

        open_btn = tk.Label(row, text=" ⊞",
                            bg=T.BG_PANEL, fg="#0284c7",
                            font=(T.FONT_MONO, T.FONT_SIZE_SMALL), cursor="hand2")
        open_btn.pack(side=tk.LEFT)
        open_btn.bind("<Button-1>", lambda e, r=rg: self._open_region_detail(r))

        del_btn = tk.Label(row, text=" x",
                           bg=T.BG_PANEL, fg="#ef4444",
                           font=(T.FONT_MONO, T.FONT_SIZE_SMALL), cursor="hand2")
        del_btn.pack(side=tk.LEFT)
        del_btn.bind("<Button-1>", lambda e, r=rg: self._remove_region(r))

        return dict(inner=inner, swatch=swatch, lbl=lbl, vis_btn=vis)

    def _region_label_text(self, rg: Region) -> str:
        if not self.times.size or not self.wavenumbers.size:
            return f" [{rg.label}] "
        t_lo = self.times[rg.ti0];   t_hi = self.times[rg.ti1]
        w_lo = self.wavenumbers[min(rg.wi0, rg.wi1)]
        w_hi = self.wavenumbers[max(rg.wi0, rg.wi1)]
        return f" [{rg.label}] t:[{t_lo:.3g},{t_hi:.3g}] wn:[{w_lo:.0f},{w_hi:.0f}] "

    def _update_region_widgets(self, rg: Region):
        w = self._region_widgets.get(rg.uid)
        if w is None:
            return
        focused = rg is self._focused_region
        w["inner"].configure(bg=rg.color if focused else T.COLOR_SPINE)
        w["swatch"].configure(bg=rg.color)
        w["lbl"].configure(text=self._region_label_text(rg), fg=rg.color)
        w["vis_btn"].configure(text="O" if rg.visible else "o")

    def _pick_region_color(self, rg: Region):
        c = colorchooser.askcolor(color=rg.color, title=f"Color for {rg.label}",
                                  parent=self.root)
        if c and c[1]:
            rg.color = c[1]
            self._draw_region_rect(rg)
            self._canvas_heat.draw_idle()
            if rg.detail_window is not None:
                rg.detail_window.update_color()
            self._update_region_widgets(rg)
            self._sync_echem_markers()

    def _toggle_region_visible(self, rg: Region):
        rg.visible = not rg.visible
        self._draw_region_rect(rg)
        self._canvas_heat.draw_idle()
        self._update_region_widgets(rg)
        self._sync_echem_markers()


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

def run_analysis_window(csv_path: str):
    root = tk.Tk()
    AnalysisWindow(root, csv_path)
    root.mainloop()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python analysis_window.py <file.csv>")
        sys.exit(1)
    run_analysis_window(sys.argv[1])